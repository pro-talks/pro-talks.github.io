import io
import time
import openpyxl
import openpyxl.styles
from openpyxl.styles import Border, Side
from openpyxl.cell import WriteOnlyCell
from pyexcelerate import Workbook as PWorkbook, Border as PBorder, Color, Style, Font, Borders

def prepare_data(n_rows):
    data = {}

    rows = []
    for row_index in range(n_rows):
        rows.append(['01.01.2023', 1234, 1.234, 'Description with date 2023.01.19 17:28',
       'Address', '123-4567-89', '12', '5.00', '7.00', '20.00%'])

    data['headers'] = ['Date', 'Number1', 'Number2', 'Description', 'Location', 'Code',
                       'value1 %', 'value2 %', 'value3 %', 'value4 %']
    data['footer'] = data['headers']
    data['rows'] = rows

    return data


def create_file(file_name, file):
    with open(file_name, "wb") as f:
        f.write(file.getbuffer())

def generate_xlsx_openpyxl(data: dict) -> io.BytesIO:

    BORDER_MEDIUM = Border(
        left=Side(color='00000000', style='medium'),
        top=Side(color='00000000', style='medium'),
        right=Side(color='00000000', style='medium'),
        bottom=Side(color='00000000', style='medium'),
    )
    BORDER_THIN = Border(
        left=Side(color='00000000', style='thin'),
        top=Side(color='00000000', style='thin'),
        right=Side(color='00000000', style='thin'),
        bottom=Side(color='00000000', style='thin'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column's width
    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[letter].width = 14

    # Set style for headers
    for col, value in enumerate(data['headers'], start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM

    for row_index, row in enumerate(data['rows'], start=1):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=1 + row_index, column=col, value=value)
            cell.border = BORDER_THIN

    # Set style for footer
    row_index = 1 + len(data['rows'])
    for col, value in enumerate(data['footer'], start=1):
        cell = ws.cell(row=row_index, column=col, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM

    output = io.BytesIO()
    wb.save(output)

    return output

def generate_xlsx_openpyxl_optimization_mode(data: dict) -> io.BytesIO:

    BORDER_MEDIUM = Border(
        left=Side(color='00000000', style='medium'),
        top=Side(color='00000000', style='medium'),
        right=Side(color='00000000', style='medium'),
        bottom=Side(color='00000000', style='medium'),
    )
    BORDER_THIN = Border(
        left=Side(color='00000000', style='thin'),
        top=Side(color='00000000', style='thin'),
        right=Side(color='00000000', style='thin'),
        bottom=Side(color='00000000', style='thin'),
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[letter].width = 14

    headers = []
    for col, value in enumerate(data['headers'], start=1):
        cell = WriteOnlyCell(ws, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM
        headers.append(cell)
    ws.append(headers)

    for row_index, row in enumerate(data['rows'], start=1):
        write_row = []
        for col, value in enumerate(row, start=1):
            cell = WriteOnlyCell(ws, value=value)
            cell.border = BORDER_THIN
            write_row.append(cell)
        ws.append(write_row)

    footer = []
    for col, value in enumerate(data['footer'], start=1):
        cell = WriteOnlyCell(ws, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM
        footer.append(cell)
    ws.append(footer)

    output = io.BytesIO()
    wb.save(output)

    return output

def generate_xlsx_openpyxl_optimization_mode_plus_defaul_border(data: dict) -> io.BytesIO:

    BORDER = Border()

    BORDER_MEDIUM = Border(
        left=Side(color='00000000', style='medium'),
        top=Side(color='00000000', style='medium'),
        right=Side(color='00000000', style='medium'),
        bottom=Side(color='00000000', style='medium'),
    )
    BORDER_THIN = Border(
        left=Side(color='00000000', style='thin'),
        top=Side(color='00000000', style='thin'),
        right=Side(color='00000000', style='thin'),
        bottom=Side(color='00000000', style='thin'),
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[letter].width = 14

    # add default border
    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[letter].border = BORDER_THIN

    headers = []
    for col, value in enumerate(data['headers'], start=1):
        cell = WriteOnlyCell(ws, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM
        headers.append(cell)
    ws.append(headers)

    for row_index, row in enumerate(data['rows'], start=1):
        ws.append(row)

    footer = []
    for col, value in enumerate(data['footer'], start=1):
        cell = WriteOnlyCell(ws, value=value)
        cell.style = 'Headline 4'
        cell.border = BORDER_MEDIUM
        footer.append(cell)
    ws.append(footer)

    output = io.BytesIO()
    wb.save(output)

    return output

def generate_xlsx_pyexcelerate(data: dict) -> io.BytesIO:

    BORDER_MEDIUM = Borders.Borders(
        top = PBorder.Border(color=Color.BLACK, style='medium'),
        left = PBorder.Border(color=Color.BLACK, style='medium'),
        right = PBorder.Border(color=Color.BLACK, style='medium'),
        bottom = PBorder.Border(color=Color.BLACK, style='medium')
    )

    BORDER_THIN = Borders.Borders(
        top = PBorder.Border(color=Color.BLACK, style='thin'),
        left = PBorder.Border(color=Color.BLACK, style='thin'),
        right = PBorder.Border(color=Color.BLACK, style='thin'),
        bottom = PBorder.Border(color=Color.BLACK, style='thin')
    )

    HEADERS = Style(
        borders=BORDER_MEDIUM,
        font=Font(color=Color(31, 73, 125, 255), bold=True)
    )

    CELL = Style(
        borders=BORDER_THIN
    )

    wb = PWorkbook()
    ws = wb.new_sheet('Sheet')
    for col in range(1, 8):
        ws.set_col_style(col, Style(size=14))

    for col, value in enumerate(data['headers'], start=1):
        ws.set_cell_value(1, col, value)
        ws.set_cell_style(1, col, HEADERS)

    for row_index, row in enumerate(data['rows'], start=1):
        for col, value in enumerate(row, start=1):
            ws.set_cell_value(1 + row_index, col, value)
            ws.set_cell_style(1 + row_index, col, CELL)

    row_index = len(data['rows'])
    for col, value in enumerate(data['footer'], start=1):
        ws.set_cell_value(1 + row_index, col, value)
        ws.set_cell_style(1 + row_index, col, HEADERS)

    output = io.BytesIO()
    wb.save(output)

    return output


def generate_xlsx_pyexcelerate_optimization_mode(data: dict) -> io.BytesIO:

    BORDER_MEDIUM = Borders.Borders(
        top=PBorder.Border(color=Color.BLACK, style='medium'),
        left=PBorder.Border(color=Color.BLACK, style='medium'),
        right=PBorder.Border(color=Color.BLACK, style='medium'),
        bottom=PBorder.Border(color=Color.BLACK, style='medium')
    )

    BORDER_THIN = Borders.Borders(
        top=PBorder.Border(color=Color.BLACK, style='thin'),
        left=PBorder.Border(color=Color.BLACK, style='thin'),
        right=PBorder.Border(color=Color.BLACK, style='thin'),
        bottom=PBorder.Border(color=Color.BLACK, style='thin')
    )

    HEADERS = Style(
        borders=BORDER_MEDIUM,
        font=Font(color=Color(31, 73, 125, 255), bold=True)
    )

    CELL = Style(
        borders=BORDER_THIN
    )

    arr = [data['headers']] + data['rows'] + [data['footer']]

    wb = PWorkbook()
    ws = wb.new_sheet('Sheet', arr)

    # Set styles
    for col in range(1, 8):
        ws.set_col_style(col, Style(size=14))

    for col, value in enumerate(data['headers'], start=1):
        ws.set_cell_style(1, col, HEADERS)

    for row_index, row in enumerate(data['rows'], start=1):
        for col, value in enumerate(row, start=1):
            ws.set_cell_style(1 + row_index, col, CELL)

    row_index = len(data['rows'])
    for col, value in enumerate(data['footer'], start=1):
        ws.set_cell_style(1 + row_index + 1, col, HEADERS)

    output = io.BytesIO()
    wb.save(output)

    return output

def pyexcelerate_optimization_mode_ignore_borders_if_file_is_huge(data: dict) -> io.BytesIO:

    BORDER_MEDIUM = Borders.Borders(
        top=PBorder.Border(color=Color.BLACK, style='medium'),
        left=PBorder.Border(color=Color.BLACK, style='medium'),
        right=PBorder.Border(color=Color.BLACK, style='medium'),
        bottom=PBorder.Border(color=Color.BLACK, style='medium')
    )

    BORDER_THIN = Borders.Borders(
        top=PBorder.Border(color=Color.BLACK, style='thin'),
        left=PBorder.Border(color=Color.BLACK, style='thin'),
        right=PBorder.Border(color=Color.BLACK, style='thin'),
        bottom=PBorder.Border(color=Color.BLACK, style='thin')
    )

    HEADERS = Style(
        borders=BORDER_MEDIUM,
        font=Font(color=Color(31, 73, 125, 255), bold=True)
    )

    CELL = Style(
        borders=BORDER_THIN
    )

    arr = [data['headers']] + data['rows'] + [data['footer']]

    wb = PWorkbook()
    ws = wb.new_sheet('Sheet', arr)

    # Set styles
    for col in range(1, 8):
        ws.set_col_style(col, Style(size=14))

    for col, value in enumerate(data['headers'], start=1):
        ws.set_cell_style(1, col, HEADERS)

    # setting styles costs when file is huge
    big_file = len(data['rows']) > 1000

    if not big_file:
        for row_index, row in enumerate(data['rows'], start=1):
            for col, value in enumerate(row, start=1):
                ws.set_cell_style(1 + row_index, col, CELL)

    row_index = len(data['rows'])
    for col, value in enumerate(data['footer'], start=1):
        ws.set_cell_style(1 + row_index + 1, col, HEADERS)

    output = io.BytesIO()
    wb.save(output)

    return output



n_rows = 10
data = prepare_data(n_rows)


start = time.time()
file = generate_xlsx_openpyxl(data)
create_file("openpyxl.xlsx", file)
print('openpyxl, time is ' + str(time.time() - start))

start = time.time()
file = generate_xlsx_openpyxl_optimization_mode(data)
create_file("openpyxl_optimization_mode.xlsx", file)
print('openpyxl_optimization_mode, time is ' + str(time.time() - start))

start = time.time()
file = generate_xlsx_openpyxl_optimization_mode_plus_defaul_border(data)
create_file("openpyxl_optimization_mode_plus_defaul_border.xlsx", file)
print('openpyxl_optimization_mode_plus_defaul_border, time is ' + str(time.time() - start))

start = time.time()
file = generate_xlsx_pyexcelerate(data)
create_file("pyexcelerate.xlsx", file)
print('pyexcelerate, time is ' + str(time.time() - start))

start = time.time()
file = generate_xlsx_pyexcelerate_optimization_mode(data)
create_file("pyexcelerate_optimization_mode.xlsx", file)
print('pyexcelerate_optimization_mode, time is ' + str(time.time() - start))

start = time.time()
file = pyexcelerate_optimization_mode_ignore_borders_if_file_is_huge(data)
create_file("pyexcelerate_optimization_mode_ignore_borders_if_file_is_huge.xlsx", file)
print('pyexcelerate_optimization_mode_ignore_borders_if_file_is_huge, time is ' + str(time.time() - start))